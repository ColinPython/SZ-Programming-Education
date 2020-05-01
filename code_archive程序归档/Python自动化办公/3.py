# -*- coding: utf-8 -*-
# author: ColinPython_榴莲老师
# 关注微信公众号：“ 青联科创 ”获得等多有趣代码教程
# 日期:2020-5-1,使用的工具:PyCharm,文件名:3

from openpyxl import Workbook  # 导入工作本工具库
from openpyxl.utils import get_column_letter  # 导入库，获得列的字母名称 将列索引转换为列字母

wb = Workbook()   # 创建工作本
dest_filename = '文件名.xlsx'  # 设置文件的名字 并设置文本后缀名格式
ws1 = wb.active  # 设置成主要工作表单​
ws1.title = "第一个表单"  # 表单名称
for row in range(1, 40):  # 使用循环语句生成   range(行)
    ws1.append(range(600))  # range(列)

ws2 = wb.create_sheet(title="第二个表单")  # 创建第二个工作表单
ws2['F5'] = 3.14  # 通过列表索引的方式 在F列 第五行 赋值输入3.14

ws3 = wb.create_sheet(title="第三个表单")  # 创建第三个工作表单
for row in range(10, 20):  # 取第10行到19行  作为编辑区域
    for col in range(27, 54):  # 取第27列 到 53列 作为编辑区域
        _ = ws3.cell(
            column=col,  # 列
            row=row,  # 行
            value="{0}".format(  # 填写的值  format 格式化
                get_column_letter(col)))

print(ws3['AA10'].value)  # 在控制面版 显示 AA列 第10行的值（内容）

wb.save(filename=dest_filename)  # 将工作表保存 （保存的文件名称为目标文件的名称）
